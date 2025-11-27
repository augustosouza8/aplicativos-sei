#!/usr/bin/env python3
"""Script para gerar planilha Excel de exemplo com processos enriquecidos."""

import sys
from pathlib import Path

# Adicionar src ao path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from sei_client.client import SeiClient
from sei_client.config import load_settings
from sei_client.models import FilterOptions, PaginationOptions, EnrichmentOptions
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def criar_planilha_detalhada(processos, caminho_saida: Path):
    """Cria planilha Excel detalhada com processos e documentos."""
    wb = Workbook()
    
    # Remover planilha padrão se existir
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # ===== ABA 1: PROCESSOS =====
    ws_processos = wb.create_sheet("Processos")
    
    # Cabeçalho dos processos
    cabecalho_processos = [
        "Número do Processo",
        "ID Procedimento",
        "Categoria",
        "Título",
        "Tipo/Especificidade",
        "Responsável",
        "CPF Responsável",
        "Visualizado",
        "Marcadores",
        "Documentos Novos",
        "Anotações",
        "Sigiloso",
        "Quantidade de Documentos",
        "Assinantes",
        "URL",
        "Hash",
    ]
    
    # Estilizar cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, cabecalho in enumerate(cabecalho_processos, 1):
        cell = ws_processos.cell(row=1, column=col_idx, value=cabecalho)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados dos processos
    for row_idx, proc in enumerate(processos, 2):
        ws_processos.cell(row=row_idx, column=1, value=proc.numero_processo)
        ws_processos.cell(row=row_idx, column=2, value=proc.id_procedimento)
        ws_processos.cell(row=row_idx, column=3, value=proc.categoria)
        ws_processos.cell(row=row_idx, column=4, value=proc.titulo or "")
        ws_processos.cell(row=row_idx, column=5, value=proc.tipo_especificidade or "")
        ws_processos.cell(row=row_idx, column=6, value=proc.responsavel_nome or "")
        ws_processos.cell(row=row_idx, column=7, value=proc.responsavel_cpf or "")
        ws_processos.cell(row=row_idx, column=8, value="Sim" if proc.visualizado else "Não")
        ws_processos.cell(row=row_idx, column=9, value=", ".join(proc.marcadores) if proc.marcadores else "")
        ws_processos.cell(row=row_idx, column=10, value="Sim" if proc.tem_documentos_novos else "Não")
        ws_processos.cell(row=row_idx, column=11, value="Sim" if proc.tem_anotacoes else "Não")
        ws_processos.cell(row=row_idx, column=12, value="Sim" if proc.eh_sigiloso else "Não")
        ws_processos.cell(row=row_idx, column=13, value=len(proc.documentos))
        ws_processos.cell(row=row_idx, column=14, value=", ".join(proc.assinantes) if proc.assinantes else "")
        ws_processos.cell(row=row_idx, column=15, value=proc.url or "")
        ws_processos.cell(row=row_idx, column=16, value=proc.hash or "")
    
    # Ajustar largura das colunas
    col_widths = [25, 15, 12, 40, 25, 30, 15, 12, 30, 15, 12, 10, 18, 40, 50, 15]
    for col_idx, width in enumerate(col_widths, 1):
        ws_processos.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ===== ABA 2: DOCUMENTOS =====
    ws_documentos = wb.create_sheet("Documentos")
    
    cabecalho_documentos = [
        "Processo",
        "ID Documento",
        "Título",
        "Tipo",
        "É Novo",
        "É Sigiloso",
        "URL Visualização",
        "URL Download",
        "Assinantes",
        "Indicadores",
        "Número Documento",
        "Metadados (JSON)",
    ]
    
    # Estilizar cabeçalho
    for col_idx, cabecalho in enumerate(cabecalho_documentos, 1):
        cell = ws_documentos.cell(row=1, column=col_idx, value=cabecalho)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados dos documentos
    row_idx = 2
    for proc in processos:
        for doc in proc.documentos:
            ws_documentos.cell(row=row_idx, column=1, value=proc.numero_processo)
            ws_documentos.cell(row=row_idx, column=2, value=doc.id_documento or "")
            ws_documentos.cell(row=row_idx, column=3, value=doc.titulo or "")
            ws_documentos.cell(row=row_idx, column=4, value=doc.tipo or "")
            ws_documentos.cell(row=row_idx, column=5, value="Sim" if doc.eh_novo else "Não")
            ws_documentos.cell(row=row_idx, column=6, value="Sim" if doc.eh_sigiloso else "Não")
            ws_documentos.cell(row=row_idx, column=7, value=doc.visualizacao_url or "")
            ws_documentos.cell(row=row_idx, column=8, value=doc.download_url or "")
            ws_documentos.cell(row=row_idx, column=9, value=", ".join(doc.assinantes) if doc.assinantes else "")
            ws_documentos.cell(row=row_idx, column=10, value=", ".join(doc.indicadores) if doc.indicadores else "")
            ws_documentos.cell(row=row_idx, column=11, value=doc.metadados.get("numero_documento", ""))
            
            # Metadados como JSON (para ver todos os campos extras)
            import json
            metadados_str = json.dumps(doc.metadados, ensure_ascii=False, indent=2) if doc.metadados else ""
            ws_documentos.cell(row=row_idx, column=12, value=metadados_str)
            
            row_idx += 1
    
    # Ajustar largura das colunas de documentos
    col_widths_doc = [25, 15, 40, 30, 10, 12, 50, 50, 40, 30, 20, 60]
    for col_idx, width in enumerate(col_widths_doc, 1):
        ws_documentos.column_dimensions[get_column_letter(col_idx)].width = width
    
    # ===== ABA 3: RESUMO =====
    ws_resumo = wb.create_sheet("Resumo")
    
    total_processos = len(processos)
    total_documentos = sum(len(p.documentos) for p in processos)
    processos_recebidos = sum(1 for p in processos if p.categoria == "Recebidos")
    processos_gerados = sum(1 for p in processos if p.categoria == "Gerados")
    
    resumo_data = [
        ["RESUMO DOS DADOS COLETADOS", ""],
        ["", ""],
        ["Total de Processos", total_processos],
        ["  - Recebidos", processos_recebidos],
        ["  - Gerados", processos_gerados],
        ["", ""],
        ["Total de Documentos", total_documentos],
        ["Média de Documentos por Processo", f"{total_documentos / total_processos:.1f}" if total_processos > 0 else "0"],
        ["", ""],
        ["Processos com Documentos Novos", sum(1 for p in processos if p.tem_documentos_novos)],
        ["Processos com Anotações", sum(1 for p in processos if p.tem_anotacoes)],
        ["Processos Sigilosos", sum(1 for p in processos if p.eh_sigiloso)],
        ["", ""],
        ["NOTA:", "Esta planilha mostra todos os dados coletados após o enriquecimento."],
        ["", "Os processos foram abertos individualmente e seus documentos foram extraídos."],
    ]
    
    for row_idx, row_data in enumerate(resumo_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_resumo.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Título
                cell.font = Font(bold=True, size=14)
    
    ws_resumo.column_dimensions["A"].width = 40
    ws_resumo.column_dimensions["B"].width = 30
    
    # Salvar
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    return caminho_saida

def main():
    print("=" * 70)
    print("GERACAO DE PLANILHA EXEMPLO - PROCESSOS ENRIQUECIDOS")
    print("=" * 70)
    
    settings = load_settings()
    client = SeiClient(settings=settings)
    
    try:
        print("\n[1/4] Fazendo login no SEI...")
        client.login()
        print(f"[OK] Login bem-sucedido. Unidade: {settings.unidade_alvo}")
        
        print("\n[2/4] Coletando 5 processos...")
        filtros = FilterOptions()
        paginacao = PaginationOptions()
        
        todos_processos, processos_filtrados = client.collect_processes(filtros, paginacao)
        
        # Limitar a 5 processos para o exemplo
        processos_para_enriquecer = todos_processos[:5]
        
        print(f"[OK] {len(processos_para_enriquecer)} processo(s) selecionado(s) para enriquecimento")
        for i, proc in enumerate(processos_para_enriquecer, 1):
            print(f"  {i}. {proc.numero_processo} ({proc.categoria})")
        
        print("\n[3/4] Enriquecendo processos (abrindo e coletando documentos)...")
        print("      Isso pode levar alguns minutos...")
        
        enrichment = EnrichmentOptions(coletar_documentos=True)
        processos_enriquecidos = client.enrich_processes(processos_para_enriquecer, enrichment)
        
        total_documentos = sum(len(p.documentos) for p in processos_enriquecidos)
        print(f"[OK] Enriquecimento concluído!")
        print(f"      - Processos processados: {len(processos_enriquecidos)}")
        print(f"      - Total de documentos coletados: {total_documentos}")
        
        print("\n[4/4] Gerando planilha Excel detalhada...")
        caminho_saida = Path("saida") / "exemplo_processos_enriquecidos.xlsx"
        caminho_final = criar_planilha_detalhada(processos_enriquecidos, caminho_saida)
        
        print(f"\n[OK] Planilha gerada com sucesso!")
        print(f"      Arquivo: {caminho_final}")
        print(f"\n      A planilha contém 3 abas:")
        print(f"      1. 'Processos' - Dados completos de cada processo")
        print(f"      2. 'Documentos' - Lista detalhada de todos os documentos")
        print(f"      3. 'Resumo' - Estatísticas gerais")
        print(f"\n" + "=" * 70)
        print("Concluído!")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n[ERRO] Erro: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1
    finally:
        client.close()
    
    return 0

if __name__ == "__main__":
    sys.exit(main())

